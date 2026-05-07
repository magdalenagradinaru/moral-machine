import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from io import BytesIO


def load_data(scenarii_path: str, human_path: str, llm_path: str) -> pd.DataFrame:
    """Load and merge all data into a single DataFrame."""
    scenarii = pd.read_csv(scenarii_path)
    
    try:
        human = pd.read_csv(human_path)
    except FileNotFoundError:
        human = pd.DataFrame(columns=["id", "human_decision", "human_reasoning"])
    
    try:
        llm = pd.read_csv(llm_path)
    except FileNotFoundError:
        llm = pd.DataFrame(columns=["id", "llm_decision", "llm_reasoning", "llm_moral_principle"])
    
    df = scenarii.copy()
    
    if not human.empty:
        df = df.merge(human[["id", "human_decision", "human_reasoning"]], on="id", how="left")
    else:
        df["human_decision"] = None
        df["human_reasoning"] = None
    
    if not llm.empty:
        df = df.merge(llm[["id", "llm_decision", "llm_reasoning", "llm_moral_principle"]], on="id", how="left")
    else:
        df["llm_decision"] = None
        df["llm_reasoning"] = None
        df["llm_moral_principle"] = None
    
    # Calculate match
    if "human_decision" in df.columns and "llm_decision" in df.columns:
        df["match"] = df.apply(
            lambda row: "✅ DA" if (
                pd.notna(row.get("human_decision")) and
                pd.notna(row.get("llm_decision")) and
                str(row["human_decision"]).strip().upper() == str(row["llm_decision"]).strip().upper()
            ) else ("❌ NU" if (pd.notna(row.get("human_decision")) and pd.notna(row.get("llm_decision"))) else "—"),
            axis=1
        )
    
    return df


def compute_stats(df: pd.DataFrame) -> dict:
    """Compute summary statistics."""
    stats = {}
    
    completed = df[df["match"] != "—"]
    if len(completed) > 0:
        matches = (completed["match"] == "✅ DA").sum()
        stats["concordanta_pct"] = round(matches / len(completed) * 100, 1)
        stats["concordanta_count"] = int(matches)
        stats["divergenta_count"] = int(len(completed) - matches)
        stats["total_comparate"] = int(len(completed))
    else:
        stats["concordanta_pct"] = 0
        stats["concordanta_count"] = 0
        stats["divergenta_count"] = 0
        stats["total_comparate"] = 0
    
    # Most common moral principles
    if "llm_moral_principle" in df.columns:
        principles = df["llm_moral_principle"].dropna()
        stats["principii"] = principles.value_counts().to_dict()
    
    # Human decisions distribution
    if "human_decision" in df.columns:
        human_dec = df["human_decision"].dropna()
        stats["human_A"] = int((human_dec.str.upper() == "A").sum())
        stats["human_B"] = int((human_dec.str.upper() == "B").sum())
    
    # LLM decisions distribution
    if "llm_decision" in df.columns:
        llm_dec = df["llm_decision"].dropna()
        stats["llm_A"] = int((llm_dec.str.upper() == "A").sum())
        stats["llm_B"] = int((llm_dec.str.upper() == "B").sum())
    
    return stats


def fig_concordanta_pie(stats: dict) -> go.Figure:
    """Pie chart for concordance."""
    labels = ["Decizii identice", "Decizii diferite"]
    values = [stats.get("concordanta_count", 0), stats.get("divergenta_count", 0)]
    colors = ["#22c55e", "#ef4444"]
    
    fig = go.Figure(data=[go.Pie(
        labels=labels,
        values=values,
        hole=0.5,
        marker_colors=colors,
        textinfo="label+percent",
        textfont_size=14,
    )])
    fig.update_layout(
        title=dict(text="Concordanță Om vs. AI", font=dict(size=18)),
        showlegend=True,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="#f1f5f9"),
        margin=dict(t=50, b=20, l=20, r=20),
        height=350
    )
    return fig


def fig_principii_bar(stats: dict) -> go.Figure:
    """Bar chart for moral principles used by LLM."""
    principii = stats.get("principii", {})
    if not principii:
        return go.Figure()
    
    sorted_p = sorted(principii.items(), key=lambda x: x[1], reverse=True)
    labels = [p[0] for p in sorted_p]
    values = [p[1] for p in sorted_p]
    
    colors = px.colors.qualitative.Set3[:len(labels)]
    
    fig = go.Figure(data=[go.Bar(
        x=values,
        y=labels,
        orientation="h",
        marker_color=colors,
        text=values,
        textposition="outside",
        textfont=dict(size=13)
    )])
    fig.update_layout(
        title=dict(text="Principii Morale Folosite de AI", font=dict(size=18)),
        xaxis_title="Număr de scenarii",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="#f1f5f9"),
        xaxis=dict(gridcolor="#334155"),
        yaxis=dict(autorange="reversed"),
        margin=dict(t=50, b=30, l=10, r=60),
        height=max(300, len(labels) * 55)
    )
    return fig


def fig_decizii_grouped(stats: dict) -> go.Figure:
    """Grouped bar chart comparing A/B decisions between human and LLM."""
    fig = go.Figure(data=[
        go.Bar(
            name="Om",
            x=["Opțiunea A", "Opțiunea B"],
            y=[stats.get("human_A", 0), stats.get("human_B", 0)],
            marker_color=["#3b82f6", "#8b5cf6"],
            text=[stats.get("human_A", 0), stats.get("human_B", 0)],
            textposition="outside"
        ),
        go.Bar(
            name="AI (LLM)",
            x=["Opțiunea A", "Opțiunea B"],
            y=[stats.get("llm_A", 0), stats.get("llm_B", 0)],
            marker_color=["#06b6d4", "#f59e0b"],
            text=[stats.get("llm_A", 0), stats.get("llm_B", 0)],
            textposition="outside"
        )
    ])
    fig.update_layout(
        title=dict(text="Distribuția Deciziilor A vs. B", font=dict(size=18)),
        barmode="group",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="#f1f5f9"),
        yaxis=dict(gridcolor="#334155"),
        legend=dict(bgcolor="rgba(0,0,0,0)"),
        margin=dict(t=50, b=30, l=30, r=30),
        height=350
    )
    return fig


def fig_match_by_category(df: pd.DataFrame) -> go.Figure:
    """Concordance rate by category."""
    if "categorie" not in df.columns or "match" not in df.columns:
        return go.Figure()
    
    completed = df[df["match"] != "—"].copy()
    if completed.empty:
        return go.Figure()
    
    completed["is_match"] = completed["match"] == "✅ DA"
    cat_stats = completed.groupby("categorie")["is_match"].agg(["sum", "count"]).reset_index()
    cat_stats["pct"] = (cat_stats["sum"] / cat_stats["count"] * 100).round(1)
    cat_stats = cat_stats.sort_values("pct", ascending=True)
    
    colors = ["#22c55e" if p >= 50 else "#ef4444" for p in cat_stats["pct"]]
    
    fig = go.Figure(data=[go.Bar(
        x=cat_stats["pct"],
        y=cat_stats["categorie"],
        orientation="h",
        marker_color=colors,
        text=[f"{p}%" for p in cat_stats["pct"]],
        textposition="outside"
    )])
    fig.update_layout(
        title=dict(text="Concordanță pe Categorii (%)", font=dict(size=18)),
        xaxis=dict(range=[0, 110], gridcolor="#334155"),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="#f1f5f9"),
        margin=dict(t=50, b=30, l=10, r=60),
        height=max(300, len(cat_stats) * 55)
    )
    return fig


def export_to_excel(df: pd.DataFrame) -> BytesIO:
    """Export the full comparison table to Excel."""
    output = BytesIO()
    
    export_cols = [
        "id", "titlu", "categorie", "descriere",
        "optiunea_A", "optiunea_B",
        "human_decision", "human_reasoning",
        "llm_decision", "llm_reasoning", "llm_moral_principle",
        "match"
    ]
    
    available = [c for c in export_cols if c in df.columns]
    export_df = df[available].copy()
    
    col_rename = {
        "id": "ID",
        "titlu": "Titlu Scenariu",
        "categorie": "Categorie",
        "descriere": "Descriere",
        "optiunea_A": "Opțiunea A",
        "optiunea_B": "Opțiunea B",
        "human_decision": "Decizie Om",
        "human_reasoning": "Raționament Om",
        "llm_decision": "Decizie AI",
        "llm_reasoning": "Raționament AI",
        "llm_moral_principle": "Principiu Moral AI",
        "match": "Concordanță"
    }
    export_df = export_df.rename(columns={k: v for k, v in col_rename.items() if k in export_df.columns})
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Comparație Om vs AI")
        
        workbook = writer.book
        worksheet = writer.sheets["Comparație Om vs AI"]
        
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Auto-width columns
        for col_idx, col in enumerate(export_df.columns, 1):
            max_length = max(
                len(str(col)),
                export_df.iloc[:, col_idx - 1].astype(str).str.len().max() if not export_df.empty else 0
            )
            adjusted = min(max_length + 4, 60)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted
        
        # Row height
        for row in worksheet.iter_rows(min_row=2):
            worksheet.row_dimensions[row[0].row].height = 80
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        worksheet.row_dimensions[1].height = 30
    
    output.seek(0)
    return output
